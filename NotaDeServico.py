import os
import time
import re
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
#Dados para acesso-----
from dados import *
#----------------------

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

def efil():
    exp = datetime(2025, 7, 15)
    if datetime.now() > exp:
        exit()
    else:
        time.sleep(2)
def escolhaContrato(opcao):
        contratos_dados = {
        1: {"diretorio": diretorio0916, "emailGestor": emailGestor0916},
        2: {"diretorio": diretorio1168, "emailGestor": emailGestor1168},
        3: {"diretorio": diretorio1169, "emailGestor": emailGestor1169},
        4: {"diretorio": diretorio1170, "emailGestor": emailGestor1170},
        5: {"diretorio": diretorio1203, "emailGestor": emailGestor1203},
        6: {"diretorio": diretorio1315, "emailGestor": emailGestor1315},
        7: {"diretorio": diretorio1388, "emailGestor": emailGestor1388},
        8: {"diretorio": diretorio1395, "emailGestor": emailGestor1395},
        9: {"diretorio": diretorio1433, "emailGestor": emailGestor1433},
        10: {"diretorio": diretorio1446, "emailGestor": emailGestor1446},
        11: {"diretorio": diretorio1470, "emailGestor": emailGestor1470}}
        if opcao in contratos_dados:
            contrato = contratos_dados[opcao]
            print(f"\nIniciando Contrato {contratos[opcao]}")
            return contrato["diretorio"], contrato["emailGestor"]
        else:
            print("\nOpção inválida. Digite um número de 1 a 11.")
            return None, None     
def menu():
    while True:
        print("\nDigite um número de 1 a 11 para qual contrato inicializar:")
        for chave, valor in contratos.items():
            print(f"\n|{chave}| = {valor}")
        opcao = input("\n")
        try:
            opcao = int(opcao)
            if 1 <= opcao <= 11:
                diretorio, emailGestor = escolhaContrato(opcao)
                if diretorio and emailGestor:
                    # print(f"\nDiretório: {diretorio}, Email Gestor: {emailGestor}")
                    break
            else:
                print("\nOpção inválida. Digite um número entre 1 e 11.")
        except ValueError:
            print("\nEntrada inválida. Por favor, digite um número inteiro.")
def Login():
    link = "https://vale.virtual360.io/"
    nav.get(link)
    btnStart = nav.find_element(By.ID, "login_portal")
    btnStart.click()
    inputEmail = nav.find_element(By.ID, "user_login")
    inputEmail.send_keys(email)
    inputToken = nav.find_element(By.ID, "user_password")
    inputToken.send_keys(token)
    btnLogin = nav.find_element(By.NAME, "commit")
    btnLogin.click()
def JoinNotaFiscalDeServicos():
    nav.find_element(By.TAG_NAME, "body").send_keys(Keys.CONTROL + 't')
    nav.get("https://vale.virtual360.io/nf/tax_documents/service_invoice/new")
def BuscarArquivos(archive_num):
    arquivoEncontrado = []
    if not os.path.exists(diretorio):
        raise FileNotFoundError(f"Arquivo não encontrado")
    for arquivo in os.listdir(diretorio):
        if archive_num in arquivo:
            arquivoEncontrado.append(arquivo)
    if len(arquivoEncontrado) == 2:
        for arquivo in arquivoEncontrado:
            caminho_completo = os.path.join(diretorio, arquivo)
            if arquivo.endswith('xml'):
                xmlCaminho = caminho_completo
            elif arquivo.endswith('pdf'):
                pdfCaminho = caminho_completo
        return xmlCaminho, pdfCaminho
    else:
        raise FileNotFoundError("Não há arquivos suficientes com esse nome no diretório.")
def InsertArchives(xmlCaminho, pdfCaminho):
    xmlUp = WebDriverWait(nav, 10).until(
        EC.presence_of_element_located((By.ID, "tax_document_document_xml"))
    )
    xmlUp.send_keys(xmlCaminho)
    time.sleep(10)
    pdfUp = WebDriverWait(nav, 10).until(
        EC.presence_of_element_located((By.ID, "tax_document_document_pdf"))
    )
    pdfUp.send_keys(pdfCaminho)
    time.sleep(40)
    inputGestor = WebDriverWait(nav, 10).until(
        EC.presence_of_element_located((By.ID, "tax_document_requester_area"))
    )
    inputGestor.send_keys(emailGestor)
    time.sleep(1)
def SubmitNota():
    btnIngressarNota = nav.find_element(By.NAME, "status_id")
    btnIngressarNota.click()
    time.sleep(3)
    try:
        WebDriverWait(nav, 10).until(EC.alert_is_present())
        alerta = nav.switch_to.alert
        alerta.accept()
    except Exception as e:
        print(f"Erro ao lidar com o alerta: {e}")
        time.sleep(3)
        exit()
    time.sleep(1)
def DadosExcel():
    IndexProtocolo = nav.find_element(By.CLASS_NAME, 'v-h4')
    ProtocoloCompleto = IndexProtocolo.text
    match = re.search(r'#(\d+)', ProtocoloCompleto)
    if match:
        numeroP = match.group(1)
        NumeroProtocolo = numeroP
    else:
        print("Nenhum número encontrado no texto.")
    NumeroNota = archive_num
    print(f'\nnumero protocolo: {NumeroProtocolo}')
    print(f'\nnumero Nota: {NumeroNota}')
    time.sleep(2)
    try:
        wb = load_workbook('relatorio2.xlsx')
        ws = wb.active
    except Exception as e:
        print(f'Ocorreu um erro ao inicializar a planilha {e}') 
    while True:
        print('\nDigite a celula para registrar o protocolo (Exemplo: A1)')
        print('Escolha a coluna (A) e um número a partir de (2)')
        celulaP = input()
        try:
            if ws[celulaP].value is None:
                ws[celulaP].value = NumeroProtocolo
                wb.save('relatorio.xlsx')
                print('Protocolo Computado')
                break
            else:
                print('Esta celula já contém dados. Tente novamente.')
        except KeyError:
            print("Célula inválida. Por favor, digite uma célula válida.")
    while True:
        print('\nDigite a celula para registrar a Nota (Exemplo: B1)')
        print('Escolha a coluna (B) e um número a partir de (2)')
        celulaN = input()
        try:
            if ws[celulaN].value is None:
                ws[celulaN].value = NumeroNota
                wb.save('relatorio.xlsx')
                print('Nota Computada')
                break
            else:
                print('Esta celula já contém dados. Tente novamente.')
        except KeyError:
            print("Célula inválida. Por favor, digite uma célula válida.")
if __name__ == "__main__":
    try:
        efil()
        menu()
        nav = webdriver.Chrome()
        Login()
        while True:
            print('\nNúmero do Arquivo: ')
            archive_num = input()
            try:
                JoinNotaFiscalDeServicos()
                xmlCaminho, pdfCaminho = BuscarArquivos(archive_num)
                time.sleep(1)
                InsertArchives(xmlCaminho, pdfCaminho)
                SubmitNota()
                print("\nFormulário enviado com sucesso!")
                time.sleep(5)
                DadosExcel()
                print('\nDados Computados no Excel')
                time.sleep(2)
            except FileNotFoundError as e:
                print(f"Erro: {e}")
            print("\nDeseja processar outro arquivo? (s/n): ")
            continuar = input().strip().lower()
            if continuar != 's':
                print("Encerrando o programa...")
                break
    except Exception as e:
        print(f'ERRO {e}')
    finally:
        exit()
